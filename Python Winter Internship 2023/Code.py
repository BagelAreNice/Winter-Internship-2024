# -*- coding: utf-8 -*-
"""
Created on Tue Jan  2 14:52:56 2024

@author: Andrew
"""
#issues to reslove, units Test 5
from paddleocr import PaddleOCR,draw_ocr,PPStructure,draw_structure_result,save_structure_res
import cv2 as cv
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment

table_engine = PPStructure(recovery=True, return_ocr_result_in_table=True, lang = 'en')
img = cv.imread('5.png')
img = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
# img = cv.bilateralFilter(img, 9,75,75)
# img = cv.Laplacian(img, -1, ksize=5, scale = 1, delta = 0, borderType = cv.BORDER_DEFAULT)
# thresh = cv.threshold(img,0,255,cv.THRESH_BINARY_INV | cv.THRESH_OTSU)
# img = thresh[1]
# img = cv.bitwise_not(img)

cv.imwrite('TestPrepross.png',img)






result = table_engine(img)

text = []
for element in result:
    if element['type']=='table':
        html_table = element['res']['html']
        html_data = pd.read_html(html_table)
        df = pd.DataFrame(html_data[0])
    else:
        if len(element['res']) == 0:
            pass
        else:
            text.append(element['res'][0]['text'])

#removing headers to remove a space between headers and table body        
dfnew = df.columns.to_frame().T.append(df,ignore_index=True)
dfnew.to_excel('Test1.xlsx',sheet_name='Test',engine = 'xlsxwriter',header = None)

wb = load_workbook('Test1.xlsx',data_only=True)
ws = wb.worksheets[0]

#remove indexing and header rows
ws.delete_cols(1)
cell = ws.cell(1,1)
if cell.value == None:
    ws.delete_rows(1)
    
#adding Title
counter = 1
for textLine in text[::-1]:
    if ws.cell(counter,1).value != textLine:
        ws.insert_rows(1)
        cell = ws.cell(row=1,column=1)
        cell.value = textLine
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ws.max_column)
        cell.alignment = Alignment(horizontal = 'center')
    else:
        text.pop(len(text)-counter)
    counter = counter + 1

#merging Dupe column cells
lastHeader = 'ajisdhuohrefuynfeuwco'
horizontal = 0
vertical = 0
for row in range(len(text)+1,len(dfnew.index)-len(df.index)+1+len(text)):
    for col in range(1,ws.max_column+1):
        cell = ws.cell(row=row,column=col)
        print('Row value: ' + str(row) + ' Col header: ' + str(col))
        if cell.value != None and lastHeader == cell.value:
            print(cell.value)
            #finding # of horizontal dupes
            for col2 in range(col+1,ws.max_column+1):
                print('Entered Horizontal Counter')
                #ensuring no errors
                if col2>ws.max_column+1:
                    print('Break at Col:' + str(col2) + ' Row:' + str(row))
                    break
                
                cell = ws.cell(row = row,column = col2)
                if cell.value == lastHeader:
                    horizontal = horizontal + 1
                    print('horizontal stack: ' + str(horizontal))
                else:
                    break
            
            #finding # of vertical dupes
            for row2 in range(row+1,len(dfnew.index)-len(df.index)+1+len(text)):
                print('Entered Vertical counter')
                if row2>len(dfnew.index)-len(df.index)+1+len(text):
                    print('Break at Col:' + str(col) + ' |Row:' + str(row2))
                    break
                
                cell = ws.cell(row = row2,column = col-1)
                print('Vertical loop: Cell - ' + cell.value + ' Last Header - ' + lastHeader)
                if cell.value == lastHeader:
                    vertical = vertical + 1
                    print('Vertical stack: ' + str(vertical))
                else:
                    break
            
            ws.merge_cells(start_row=row,start_column=col-1, end_row=row+vertical, end_column=col+horizontal)
            print('merge done')
            cellAlignment = ws.cell(row = row, column=col-1)
            cellAlignment.alignment = Alignment(horizontal = 'center',vertical='center')
            horizontal = 0
            vertical = 0
            
        lastHeader = cell.value
vertical = 0
#because of the dupe detection method, a single column dupe detection must be made.
for col in range(1,ws.max_column+1):
    for row in range(len(text)+1,len(dfnew.index)-len(df.index)+1+len(text)):
        print('Entering vertical dupe detection')
        cell = ws.cell(row=row,column=col)
        if cell.value != None and lastHeader == cell.value:
            if col+1>ws.max_column:
                pass
            elif ws.cell(row=row,column=col+1)==None:
                break
            
            for row2 in range(row+1,len(dfnew.index)-len(df.index)+1+len(text)):
                if row2>len(dfnew.index)-len(df.index)+1+len(text):
                    print('Break at Col:' + str(col) + ' |Row:' + str(row2))
                    break
                
                cell = ws.cell(row = row2,column = col)
                print('Vertical loop: Cell - ' + cell.value + ' Last Header - ' + lastHeader)
                if cell.value == lastHeader:
                    vertical = vertical + 1
                    print('Vertical stack: ' + str(vertical))
                else:
                    break
            ws.merge_cells(start_row=row-1,start_column=col, end_row=row+vertical, end_column=col)
            print('Vertical merge!!!!!!!!')
            cellAlignment = ws.cell(row = row-1, column=col)
            cellAlignment.alignment = Alignment(horizontal = 'center',vertical='center')
            vertical = 0
        lastHeader = cell.value

#removing unnamed cells
for col in range(1,ws.max_column+1):
    for row in range(1,ws.max_row):
        cell = ws.cell(row = row, column = col)
        if type(cell.value)==str:
            if cell.value.find('Unnamed')!=-1:
                cell.value=None


    
    
wb.save('Test1Mod.xlsx')


